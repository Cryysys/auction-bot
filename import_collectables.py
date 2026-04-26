import argparse
from pathlib import Path

from openpyxl import load_workbook

import database

RARITY_MAP = {
    "common": "common",
    "uncommon": "uncommon",
    "rare": "rare",
    "epic": "epic",
    "legendary": "legendary",
}


def _normalize_text(value: object) -> str:
    return str(value).strip() if value is not None else ""


def _normalize_rarity(value: object) -> str:
    normalized = _normalize_text(value).lower()
    return RARITY_MAP.get(normalized, normalized)


def _read_rows(excel_path: Path, sheet_name: str | None):
    workbook = load_workbook(filename=excel_path, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active

    headers = [_normalize_text(cell.value).lower() for cell in worksheet[1]]
    required = {"name", "rarity", "category", "description", "image", "release_date"}
    missing = required.difference(set(headers))
    if missing:
        raise ValueError(f"Missing required columns in Excel: {', '.join(sorted(missing))}")

    idx = {header: i for i, header in enumerate(headers)}
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        name = _normalize_text(row[idx["name"]])
        if not name:
            continue
        yield {
            "name": name,
            "rarity": _normalize_rarity(row[idx["rarity"]]),
            "category": _normalize_text(row[idx["category"]]) or "Collectibles",
            "description": _normalize_text(row[idx["description"]]),
            "image_url": _normalize_text(row[idx["image"]]),
            "release_date": _normalize_text(row[idx["release_date"]]),
        }


def main():
    parser = argparse.ArgumentParser(description="Import collectables from Excel.")
    parser.add_argument("excel_path", help="Path to source Excel file")
    parser.add_argument("--sheet", help="Optional sheet name")
    parser.add_argument(
        "--dry-run", action="store_true", help="Validate and preview without writing to DB"
    )
    args = parser.parse_args()

    excel_path = Path(args.excel_path)
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    database.init_db()

    inserted = 0
    updated = 0
    unknown_rarity = 0
    unknown_category = 0

    for row in _read_rows(excel_path, args.sheet):
        existing = database.get_collectable_by_name(row["name"])
        if row["rarity"] not in RARITY_MAP:
            unknown_rarity += 1
        if not row["category"]:
            unknown_category += 1

        if not args.dry_run:
            database.upsert_collectable(
                name=row["name"],
                category=row["category"],
                rarity=row["rarity"] or "common",
                description=row["description"],
                image_url=row["image_url"],
                release_date=row["release_date"],
                is_active=True,
            )

        if existing:
            updated += 1
        else:
            inserted += 1

    mode = "DRY RUN" if args.dry_run else "IMPORT"
    print(f"[{mode}] Collectables processed.")
    print(f"Inserted: {inserted}")
    print(f"Updated: {updated}")
    print(f"Unknown rarity values: {unknown_rarity}")
    print(f"Unknown category values: {unknown_category}")


if __name__ == "__main__":
    main()
